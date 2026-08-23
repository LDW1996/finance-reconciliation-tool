from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import re
from typing import BinaryIO

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REQUIRED_COLUMNS = {"公司代码", "分配", "原币金额"}
COMBINED_COMPANY_COLUMN = "公司代码or伙伴公司"
DISPLAY_COLUMNS = ["公司代码", "凭证编号", "分配", "文本", "原币金额"]


class ReconciliationError(Exception):
    def __init__(self, user_message: str):
        self.user_message = user_message
        super().__init__(user_message)


@dataclass
class ReconciliationResult:
    company_a_code: str
    company_b_code: str
    summary: dict[str, dict[str, float]]
    matched: pd.DataFrame
    differences: pd.DataFrame
    unmatched: pd.DataFrame
    raw_data: pd.DataFrame
    excel_bytes: bytes
    output_filename: str


def read_excel(file_obj: BinaryIO | BytesIO, filename: str) -> pd.DataFrame:
    suffix = Path(filename).suffix.lower()
    if suffix not in {".xlsx", ".xls"}:
        raise ReconciliationError("文件格式错误，请选择Excel文件")

    try:
        return pd.read_excel(file_obj)
    except Exception as exc:
        raise ReconciliationError("Excel读取失败，请检查文件是否损坏或格式是否正确") from exc


def inspect_combined_workbook(
    workbook_file: BinaryIO | BytesIO,
    filename: str,
    match_field: str = "分配",
    amount_field: str = "原币金额",
    company_field: str = COMBINED_COMPANY_COLUMN,
) -> dict[str, object]:
    report: dict[str, object] = {
        "valid": False,
        "errors": [],
        "warnings": [],
        "companyCodes": [],
        "rowCount": 0,
        "allocationCount": 0,
        "requiredColumns": [company_field, match_field, amount_field],
    }
    errors: list[str] = []
    warnings: list[str] = []

    suffix = Path(filename).suffix.lower()
    if suffix not in {".xlsx", ".xls"}:
        errors.append("文件格式错误，请选择 .xlsx 或 .xls 文件")
        report["errors"] = errors
        return report

    try:
        df = pd.read_excel(workbook_file)
    except Exception:
        errors.append("Excel读取失败，请检查文件是否损坏或格式是否正确")
        report["errors"] = errors
        return report

    report["rowCount"] = int(len(df))
    if df.empty:
        errors.append("往来表没有可分析的数据，请检查Excel内容")

    required = {company_field, match_field, amount_field}
    missing = sorted(required - set(df.columns))
    if missing:
        errors.append(f"Excel缺少字段：{'、'.join(missing)}，请检查模板")
        report["errors"] = errors
        report["warnings"] = warnings
        return report

    company_codes = sorted(
        df[company_field].dropna().astype(str).str.strip().replace("", pd.NA).dropna().unique().tolist()
    )
    report["companyCodes"] = company_codes
    if len(company_codes) != 2:
        visible_codes = "、".join(company_codes) if company_codes else "无"
        errors.append(f"公司代码or伙伴公司必须且只能包含两家公司代码，当前识别到：{visible_codes}")

    amount_values = pd.to_numeric(df[amount_field], errors="coerce")
    invalid_amounts = int(amount_values.isna().sum() - df[amount_field].isna().sum())
    if invalid_amounts > 0:
        errors.append(f"原币金额存在 {invalid_amounts} 行无法识别为数字，请检查金额列")

    normalized_allocations = _normalize_match_key(df[match_field])
    blank_allocations = int((normalized_allocations == "空白").sum())
    report["allocationCount"] = int(normalized_allocations.nunique(dropna=False))
    if blank_allocations > 0:
        warnings.append(f"存在 {blank_allocations} 行分配为空，系统会统一按“空白”分组处理")

    report["valid"] = len(errors) == 0
    report["errors"] = errors
    report["warnings"] = warnings
    return report


def validate_dataframe(df: pd.DataFrame, label: str) -> None:
    missing = sorted(REQUIRED_COLUMNS - set(df.columns))
    if missing:
        raise ReconciliationError(f"Excel缺少字段：{'、'.join(missing)}，请检查模板")
    if df.empty:
        raise ReconciliationError(f"{label}没有可分析的数据，请检查Excel内容")


def identify_company_code(df: pd.DataFrame, label: str) -> str:
    codes = df["公司代码"].dropna().astype(str).str.strip()
    codes = codes[codes != ""].unique()
    if len(codes) == 0:
        raise ReconciliationError(f"{label}无法识别公司代码，请检查公司代码字段")
    if len(codes) > 1:
        raise ReconciliationError(f"{label}包含多个公司代码，请拆分后再上传")
    return str(codes[0])


def reconcile_workbooks(
    company_a_file: BinaryIO | BytesIO,
    company_b_file: BinaryIO | BytesIO,
    company_a_filename: str,
    company_b_filename: str,
    match_field: str = "分配",
    amount_field: str = "原币金额",
) -> ReconciliationResult:
    df_a = read_excel(company_a_file, company_a_filename)
    df_b = read_excel(company_b_file, company_b_filename)
    validate_dataframe(df_a, "公司A")
    validate_dataframe(df_b, "公司B")

    for field in {match_field, amount_field}:
        if field not in df_a.columns or field not in df_b.columns:
            raise ReconciliationError(f"Excel缺少字段：{field}，请检查模板")

    company_a_code = identify_company_code(df_a, "公司A")
    company_b_code = identify_company_code(df_b, "公司B")
    if company_a_code == company_b_code:
        raise ReconciliationError("两个文件属于同一家公司，无法进行比对")

    prepared_a = _prepare(df_a, match_field, amount_field, "A")
    prepared_b = _prepare(df_b, match_field, amount_field, "B")
    comparison = prepared_a.merge(prepared_b, how="outer", on=match_field, indicator=True)

    both = comparison[comparison["_merge"] == "both"].copy()
    both["合计金额"] = both["公司A金额"].fillna(0) + both["公司B金额"].fillna(0)

    matched_rows = both[both["合计金额"].abs() < 0.000001].copy()
    difference_rows = both[both["合计金额"].abs() >= 0.000001].copy()
    unmatched_rows = comparison[comparison["_merge"] != "both"].copy()

    matched = _build_matched_sheet(matched_rows, df_a, match_field)
    differences = _build_difference_sheet(difference_rows, match_field)
    unmatched = _build_unmatched_sheet(unmatched_rows, match_field)
    raw_data = pd.concat(
        [_tag_raw(df_a, "公司A"), _tag_raw(df_b, "公司B")],
        ignore_index=True,
    )
    summary = _build_summary(matched, differences, unmatched)
    excel_bytes = _write_result_excel(summary, matched, differences, unmatched, raw_data)
    output_filename = f"{company_a_code}_{company_b_code}_对账结果.xlsx"

    return ReconciliationResult(
        company_a_code=company_a_code,
        company_b_code=company_b_code,
        summary=summary,
        matched=matched,
        differences=differences,
        unmatched=unmatched,
        raw_data=raw_data,
        excel_bytes=excel_bytes,
        output_filename=output_filename,
    )


def reconcile_combined_workbook(
    workbook_file: BinaryIO | BytesIO,
    filename: str,
    match_field: str = "分配",
    amount_field: str = "原币金额",
    company_field: str = COMBINED_COMPANY_COLUMN,
) -> ReconciliationResult:
    df = read_excel(workbook_file, filename)
    _validate_combined_dataframe(df, match_field, amount_field, company_field)

    work = df.copy()
    work[company_field] = work[company_field].astype(str).str.strip()
    company_codes = sorted(work[company_field].dropna().unique().tolist())
    if len(company_codes) != 2:
        raise ReconciliationError("往来表必须且只能包含两家公司代码，请检查公司代码or伙伴公司字段")

    company_a_code, company_b_code = company_codes
    work["_匹配键"] = _normalize_match_key(work[match_field])
    work[amount_field] = pd.to_numeric(work[amount_field], errors="coerce").fillna(0)

    grouped = (
        work.groupby(["_匹配键", company_field], dropna=False)
        .agg(
            明细行数=(amount_field, "size"),
            原币金额合计=(amount_field, "sum"),
            凭证编号=("凭证编号", _join_unique) if "凭证编号" in work.columns else (amount_field, "count"),
        )
        .reset_index()
    )
    amount_pivot = grouped.pivot(index="_匹配键", columns=company_field, values="原币金额合计").fillna(0)
    rows_pivot = grouped.pivot(index="_匹配键", columns=company_field, values="明细行数").fillna(0)

    status_by_key: dict[str, str] = {}
    totals_by_key: dict[str, dict[str, float]] = {}
    for match_key in amount_pivot.index:
        amount_a = float(amount_pivot.loc[match_key].get(company_a_code, 0))
        amount_b = float(amount_pivot.loc[match_key].get(company_b_code, 0))
        rows_a = int(rows_pivot.loc[match_key].get(company_a_code, 0))
        rows_b = int(rows_pivot.loc[match_key].get(company_b_code, 0))
        combined = amount_a + amount_b
        if rows_a and rows_b and abs(combined) < 0.000001:
            status = "匹配成功"
        elif rows_a and rows_b:
            status = "金额差异"
        else:
            status = "未匹配"
        status_by_key[match_key] = status
        totals_by_key[match_key] = {
            f"{company_a_code}原币金额合计": amount_a,
            f"{company_b_code}原币金额合计": amount_b,
            "两公司原币金额合计": combined,
            f"{company_a_code}行数": rows_a,
            f"{company_b_code}行数": rows_b,
        }

    detail = work.copy()
    detail["分配"] = detail["_匹配键"]
    detail["匹配状态"] = detail["_匹配键"].map(status_by_key)
    for column in [f"{company_a_code}原币金额合计", f"{company_b_code}原币金额合计", "两公司原币金额合计"]:
        detail[column] = detail["_匹配键"].map(lambda key: totals_by_key[key][column])
    detail["未匹配说明"] = detail.apply(
        lambda row: _unmatched_reason(row, company_field, company_a_code, company_b_code)
        if row["匹配状态"] == "未匹配"
        else "",
        axis=1,
    )
    detail = detail.drop(columns=["_匹配键"])
    trailing_columns = [
        f"{company_a_code}原币金额合计",
        f"{company_b_code}原币金额合计",
        "两公司原币金额合计",
        "匹配状态",
        "未匹配说明",
    ]
    detail = detail[[column for column in detail.columns if column not in trailing_columns] + trailing_columns]

    matched = detail[detail["匹配状态"] == "匹配成功"].copy()
    differences = detail[detail["匹配状态"] == "金额差异"].copy()
    unmatched = detail[detail["匹配状态"] == "未匹配"].copy()
    difference_summary = _build_difference_summary(
        grouped, status_by_key, company_a_code, company_b_code, company_field
    )
    summary = _build_combined_summary(matched, differences, unmatched, company_field, company_a_code, company_b_code)
    excel_bytes = _write_combined_result_excel(summary, matched, differences, unmatched, difference_summary, df)
    return ReconciliationResult(
        company_a_code=company_a_code,
        company_b_code=company_b_code,
        summary=summary,
        matched=matched,
        differences=differences,
        unmatched=unmatched,
        raw_data=df,
        excel_bytes=excel_bytes,
        output_filename=f"{company_a_code}_{company_b_code}_分配原币金额对账结果.xlsx",
    )


def _validate_combined_dataframe(
    df: pd.DataFrame,
    match_field: str,
    amount_field: str,
    company_field: str,
) -> None:
    required = {company_field, match_field, amount_field}
    missing = sorted(required - set(df.columns))
    if missing:
        raise ReconciliationError(f"Excel缺少字段：{'、'.join(missing)}，请检查模板")
    if df.empty:
        raise ReconciliationError("往来表没有可分析的数据，请检查Excel内容")


def _normalize_match_key(series: pd.Series) -> pd.Series:
    normalized = series.fillna("").astype(str).str.strip()
    return normalized.mask(normalized == "", "空白")


def _unmatched_reason(row: pd.Series, company_field: str, company_a_code: str, company_b_code: str) -> str:
    if str(row[company_field]) == company_a_code:
        return f"仅{company_a_code}存在"
    if str(row[company_field]) == company_b_code:
        return f"仅{company_b_code}存在"
    return "仅单方存在"


def _build_combined_summary(
    matched: pd.DataFrame,
    differences: pd.DataFrame,
    unmatched: pd.DataFrame,
    company_field: str,
    company_a_code: str,
    company_b_code: str,
) -> dict[str, dict[str, float]]:
    frames = {"匹配成功": matched, "金额差异": differences, "未匹配": unmatched}
    summary: dict[str, dict[str, float]] = {}
    for status, frame in frames.items():
        summary[status] = {
            "分配数量": int(frame["分配"].nunique(dropna=False)),
            "明细行数": int(len(frame)),
            f"{company_a_code}行数": int((frame[company_field].astype(str) == company_a_code).sum()),
            f"{company_b_code}行数": int((frame[company_field].astype(str) == company_b_code).sum()),
            f"{company_a_code}原币金额合计": float(frame.loc[frame[company_field].astype(str) == company_a_code, "原币金额"].sum()),
            f"{company_b_code}原币金额合计": float(frame.loc[frame[company_field].astype(str) == company_b_code, "原币金额"].sum()),
        }
    summary["合计"] = {
        "分配数量": sum(item["分配数量"] for item in summary.values()),
        "明细行数": sum(item["明细行数"] for item in summary.values()),
        f"{company_a_code}行数": sum(item[f"{company_a_code}行数"] for item in summary.values()),
        f"{company_b_code}行数": sum(item[f"{company_b_code}行数"] for item in summary.values()),
        f"{company_a_code}原币金额合计": sum(item[f"{company_a_code}原币金额合计"] for item in summary.values()),
        f"{company_b_code}原币金额合计": sum(item[f"{company_b_code}原币金额合计"] for item in summary.values()),
    }
    return summary


def _build_difference_summary(
    grouped: pd.DataFrame,
    status_by_key: dict[str, str],
    company_a_code: str,
    company_b_code: str,
    company_field: str,
) -> pd.DataFrame:
    rows = []
    for match_key, status in status_by_key.items():
        if status != "金额差异":
            continue
        subset = grouped[grouped["_匹配键"] == match_key]
        by_company = subset.set_index(company_field)
        amount_a = float(by_company["原币金额合计"].get(company_a_code, 0))
        amount_b = float(by_company["原币金额合计"].get(company_b_code, 0))
        rows.append(
            {
                "分配": match_key,
                f"{company_a_code}行数": int(by_company["明细行数"].get(company_a_code, 0)),
                f"{company_b_code}行数": int(by_company["明细行数"].get(company_b_code, 0)),
                f"{company_a_code}原币金额合计": amount_a,
                f"{company_b_code}原币金额合计": amount_b,
                "两公司金额合计(差额)": amount_a + amount_b,
                "差额绝对值": abs(amount_a + amount_b),
                f"{company_a_code}凭证编号": by_company["凭证编号"].get(company_a_code, ""),
                f"{company_b_code}凭证编号": by_company["凭证编号"].get(company_b_code, ""),
            }
        )
    return pd.DataFrame(rows)


def _write_combined_result_excel(
    summary: dict[str, dict[str, float]],
    matched: pd.DataFrame,
    differences: pd.DataFrame,
    unmatched: pd.DataFrame,
    difference_summary: pd.DataFrame,
    raw_data: pd.DataFrame,
) -> bytes:
    summary_df = pd.DataFrame([{"分类": name, **data} for name, data in summary.items()])
    matched_out = _format_export_dataframe(matched)
    differences_out = _format_export_dataframe(differences)
    unmatched_out = _format_export_dataframe(unmatched)
    raw_data_out = _format_export_dataframe(raw_data)
    difference_summary_out = _format_export_dataframe(difference_summary)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="处理汇总", index=False, startrow=2)
        matched_out.to_excel(writer, sheet_name="匹配成功", index=False)
        differences_out.to_excel(writer, sheet_name="金额差异明细", index=False)
        unmatched_out.to_excel(writer, sheet_name="未匹配", index=False)
        difference_summary_out.to_excel(writer, sheet_name="金额差异汇总", index=False)
        raw_data_out.to_excel(writer, sheet_name="原始数据", index=False)
        _write_rule_section(writer.book["处理汇总"], summary_df)
        _style_combined_workbook(writer.book, summary_df)
    return buffer.getvalue()


def _format_export_dataframe(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    if "分配" in result.columns:
        result["分配"] = result["分配"].replace({"空白": "24"})
    if "行项目" in result.columns:
        result["行项目"] = result["行项目"].map(_format_line_item)
    for column in result.columns:
        if column.endswith("凭证编号"):
            result[column] = result[column].map(_format_identifier)
        elif column in {
            "凭证编号",
            "凭证编号1",
            "凭证编号2",
            "公司代码or伙伴公司",
            "总账科目",
            "合并科目",
            "年度",
            "期间",
            "客户",
            "供应商",
        }:
            result[column] = result[column].map(_format_identifier)
    computed_columns = {"匹配状态", "未匹配说明"}
    computed_columns.update(column for column in result.columns if "原币金额合计" in str(column))
    computed_columns.add("两公司原币金额合计")
    for column in result.columns:
        if column in computed_columns:
            continue
        result[column] = result[column].where(pd.notna(result[column]), "24").replace("", "24")
    return result


def _format_identifier(value: object) -> object:
    if pd.isna(value):
        return value
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _format_line_item(value: object) -> object:
    if pd.isna(value):
        return value
    text = str(value).strip()
    if re.fullmatch(r"\d+(\.0)?", text):
        return f"{int(float(text)):06d}"
    return text


def _write_rule_section(ws, summary_df: pd.DataFrame) -> None:
    company_columns = [col for col in summary_df.columns if re.match(r"^\d+行数$", str(col))]
    company_codes = [col.removesuffix("行数") for col in company_columns[:2]]
    company_a, company_b = company_codes if len(company_codes) == 2 else ("公司A", "公司B")
    start_row = 3 + len(summary_df) + 2
    rules = [
        ("核对规则", "说明"),
        ("匹配成功", f"分配字段一致，且{company_a}与{company_b}的原币金额汇总后相加等于0。"),
        ("金额差异", f"分配字段在{company_a}、{company_b}两家公司都存在，但原币金额合计不等于0。"),
        ("未匹配", "分配字段仅存在于其中一家公司；分配为空的记录也归入未匹配。"),
        ("金额口径", f"本文件{company_a}和{company_b}的原币均为IDR，因此直接使用“原币金额”核对。"),
    ]
    for offset, row in enumerate(rules):
        ws.cell(start_row + offset, 1, row[0])
        ws.cell(start_row + offset, 2, row[1])


def _style_combined_workbook(workbook, summary_df: pd.DataFrame) -> None:
    header_fill = PatternFill("solid", fgColor="FF1F4E78")
    title_fill = PatternFill("solid", fgColor="FF17365D")
    total_fill = PatternFill("solid", fgColor="FFEAF2F8")
    rule_fill = PatternFill("solid", fgColor="FFD9EAF7")
    white_bold = Font(bold=True, color="FFFFFFFF")
    bold_font = Font(bold=True, color="FF1B1F23")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="FFD9E2DD"),
        right=Side(style="thin", color="FFD9E2DD"),
        top=Side(style="thin", color="FFD9E2DD"),
        bottom=Side(style="thin", color="FFD9E2DD"),
    )

    summary_ws = workbook["处理汇总"]
    last_summary_col = summary_ws.max_column
    summary_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_summary_col)
    title_cell = summary_ws.cell(1, 1)
    company_columns = [col for col in summary_df.columns if re.match(r"^\d+行数$", str(col))]
    company_codes = [col.removesuffix("行数") for col in company_columns[:2]]
    if len(company_codes) == 2:
        title_cell.value = f"{company_codes[0]} 与 {company_codes[1]} 分配字段 / 原币金额对账结果"
    else:
        title_cell.value = "分配字段 / 原币金额对账结果"
    title_cell.fill = title_fill
    title_cell.font = white_bold
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    summary_ws.row_dimensions[1].height = 28
    summary_ws.row_dimensions[3].height = 22
    _style_header_row(summary_ws, 3, header_fill, white_bold, header_alignment)
    _apply_table_format(summary_ws, 3, 3 + len(summary_df), thin_border)
    total_row = 3 + len(summary_df)
    for cell in summary_ws[total_row]:
        cell.fill = total_fill
        cell.font = bold_font
    rule_header_row = 3 + len(summary_df) + 2
    _style_header_row(summary_ws, rule_header_row, rule_fill, bold_font, header_alignment)
    _apply_table_format(summary_ws, rule_header_row, rule_header_row + 4, thin_border, max_col=2)
    summary_ws.freeze_panes = "A4"
    summary_ws.auto_filter.ref = f"A3:{get_column_letter(last_summary_col)}{3 + len(summary_df)}"
    _set_widths(summary_ws, {"A": 18, "B": 72, "C": 12, "D": 12, "E": 12, "F": 22, "G": 22})
    _format_numeric_columns(summary_ws, 4, 3 + len(summary_df))

    for sheet_name in ["匹配成功", "金额差异明细", "未匹配", "原始数据"]:
        ws = workbook[sheet_name]
        _style_header_row(ws, 1, header_fill, white_bold, header_alignment)
        _apply_table_format(ws, 1, ws.max_row, thin_border)
        _format_numeric_columns(ws, 2, ws.max_row)
        _auto_fit_sheet(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    diff_summary = workbook["金额差异汇总"]
    _style_header_row(diff_summary, 1, header_fill, white_bold, header_alignment)
    _apply_table_format(diff_summary, 1, diff_summary.max_row, thin_border)
    _format_numeric_columns(diff_summary, 2, diff_summary.max_row)
    _set_widths(
        diff_summary,
        {
            "A": 24,
            "B": 11,
            "C": 11,
            "D": 22,
            "E": 22,
            "F": 22,
            "G": 22,
            "H": 38,
            "I": 38,
        },
    )
    _auto_fit_sheet(diff_summary, max_width=48)
    diff_summary.freeze_panes = "A2"
    diff_summary.auto_filter.ref = diff_summary.dimensions


def _style_header_row(ws, row: int, fill: PatternFill, font: Font, alignment: Alignment) -> None:
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def _apply_table_format(ws, start_row: int, end_row: int, border: Border, max_col: int | None = None) -> None:
    max_column = max_col or ws.max_column
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, max_col=max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _format_numeric_columns(ws, start_row: int, end_row: int) -> None:
    for column_cells in ws.iter_cols(min_row=1, max_row=end_row):
        header = str(column_cells[0].value or "")
        if any(keyword in header for keyword in ["金额", "差额"]):
            for cell in column_cells[start_row - 1 :]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
        elif header.endswith("行数") or header in {"分配数量", "明细行数"}:
            for cell in column_cells[start_row - 1 :]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"


def _auto_fit_sheet(ws, min_width: float = 10, max_width: float = 42) -> None:
    for column_index in range(1, ws.max_column + 1):
        letter = get_column_letter(column_index)
        best_width = min_width
        for cell in ws[letter]:
            text = "" if cell.value is None else str(cell.value)
            best_width = max(best_width, min(max_width, len(text) * 1.15 + 3))
        ws.column_dimensions[letter].width = best_width
    for row_index in range(1, ws.max_row + 1):
        max_lines = 1
        for cell in ws[row_index]:
            if cell.value is None:
                continue
            width = ws.column_dimensions[get_column_letter(cell.column)].width or 12
            max_lines = max(max_lines, int(len(str(cell.value)) / max(width, 8)) + 1)
        ws.row_dimensions[row_index].height = min(72, max(20, max_lines * 18))


def _set_widths(ws, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _set_detail_widths(ws) -> None:
    default_widths = {
        "A": 13,
        "B": 13,
        "C": 13,
        "D": 13,
        "E": 13,
        "I": 13,
        "K": 13,
        "M": 13,
        "F": 24,
        "G": 30,
        "H": 14,
        "J": 24,
        "L": 14,
        "N": 14,
        "O": 14,
        "P": 14,
        "Q": 14,
        "R": 14,
        "S": 22,
        "T": 22,
        "U": 22,
        "V": 14,
        "W": 16,
    }
    _set_widths(ws, default_widths)


def _prepare(df: pd.DataFrame, match_field: str, amount_field: str, side: str) -> pd.DataFrame:
    work = df.copy()
    work[match_field] = work[match_field].astype(str).str.strip()
    work = work[work[match_field] != ""]
    work[amount_field] = pd.to_numeric(work[amount_field], errors="coerce").fillna(0)

    grouped = work.groupby(match_field, as_index=False).agg(
        **{
            f"公司{side}金额": (amount_field, "sum"),
            f"公司{side}凭证编号": ("凭证编号", _join_unique) if "凭证编号" in work.columns else (amount_field, "count"),
        }
    )
    return grouped


def _join_unique(values: pd.Series) -> str:
    items = [str(value).strip() for value in values.dropna().tolist() if str(value).strip()]
    unique_items = list(dict.fromkeys(items))
    return ", ".join(sorted(unique_items, key=_natural_sort_key))


def _natural_sort_key(value: str) -> list[object]:
    return [int(part) if part.isdigit() else part for part in re.split(r"(\d+)", value)]


def _build_matched_sheet(rows: pd.DataFrame, source_a: pd.DataFrame, match_field: str) -> pd.DataFrame:
    if rows.empty:
        return pd.DataFrame(columns=DISPLAY_COLUMNS + ["对方金额", "合计金额", "匹配状态"])

    base_cols = [col for col in DISPLAY_COLUMNS if col in source_a.columns]
    base = source_a[source_a[match_field].isin(rows[match_field])][base_cols].copy()
    lookup = rows.set_index(match_field)[["公司B金额", "合计金额"]]
    base = base.join(lookup, on=match_field)
    base = base.rename(columns={"公司B金额": "对方金额"})
    base["匹配状态"] = "匹配成功"
    return base


def _build_difference_sheet(rows: pd.DataFrame, match_field: str) -> pd.DataFrame:
    if rows.empty:
        return pd.DataFrame(columns=[match_field, "公司A金额", "公司B金额", "差异金额", "凭证编号"])
    result = rows[[match_field, "公司A金额", "公司B金额", "合计金额", "公司A凭证编号"]].copy()
    result = result.rename(columns={"合计金额": "差异金额", "公司A凭证编号": "凭证编号"})
    return result


def _build_unmatched_sheet(rows: pd.DataFrame, match_field: str) -> pd.DataFrame:
    if rows.empty:
        return pd.DataFrame(columns=[match_field, "金额", "未匹配原因", "凭证编号"])
    result = pd.DataFrame(
        {
            match_field: rows[match_field],
            "金额": rows["公司A金额"].fillna(rows["公司B金额"]),
            "未匹配原因": rows["_merge"].map({"left_only": "仅公司A存在", "right_only": "仅公司B存在"}),
            "凭证编号": rows["公司A凭证编号"].fillna(rows["公司B凭证编号"]),
        }
    )
    return result


def _tag_raw(df: pd.DataFrame, source: str) -> pd.DataFrame:
    result = df.copy()
    result.insert(0, "来源文件", source)
    return result


def _build_summary(
    matched: pd.DataFrame,
    differences: pd.DataFrame,
    unmatched: pd.DataFrame,
) -> dict[str, dict[str, float]]:
    matched_amount = float(matched.get("原币金额", pd.Series(dtype=float)).sum())
    difference_amount = float(differences.get("差异金额", pd.Series(dtype=float)).sum())
    unmatched_amount = float(unmatched.get("金额", pd.Series(dtype=float)).sum())
    total_count = len(matched) + len(differences) + len(unmatched)
    total_amount = matched_amount + difference_amount + unmatched_amount
    return {
        "匹配成功": {"数量": int(len(matched)), "金额": matched_amount},
        "金额差异": {"数量": int(len(differences)), "金额": difference_amount},
        "未匹配": {"数量": int(len(unmatched)), "金额": unmatched_amount},
        "合计": {"数量": int(total_count), "金额": total_amount},
    }


def _write_result_excel(
    summary: dict[str, dict[str, float]],
    matched: pd.DataFrame,
    differences: pd.DataFrame,
    unmatched: pd.DataFrame,
    raw_data: pd.DataFrame,
) -> bytes:
    summary_df = pd.DataFrame(
        [{"分类": name, "数量": data["数量"], "金额": data["金额"]} for name, data in summary.items()]
    )
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="汇总结果", index=False)
        matched.to_excel(writer, sheet_name="匹配成功", index=False)
        differences.to_excel(writer, sheet_name="金额差异", index=False)
        unmatched.to_excel(writer, sheet_name="未匹配", index=False)
        raw_data.to_excel(writer, sheet_name="原始数据", index=False)
    return buffer.getvalue()
